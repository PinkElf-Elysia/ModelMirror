"""Network-free, sealed-input compatibility facades for catalog Wave 18B."""

from __future__ import annotations

import ast
import csv
import io
import json
import math
import re
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Annotated, Any, Literal

from mcp.server.fastmcp import FastMCP
from mcp.types import ToolAnnotations
from pydantic import BaseModel, ConfigDict, Field


FileId = Annotated[
    str,
    Field(
        description="从当前受控工作区选择文件，不接受本机路径或 URI。",
        json_schema_extra={"x-modelmirror-input": "workspace-file"},
    ),
]
ArtifactName = Annotated[
    str,
    Field(
        description="新产物文件名；只写入当前工作区的可清理产物目录。",
        json_schema_extra={"x-modelmirror-input": "artifact-name"},
    ),
]

READ_ONLY = ToolAnnotations(
    readOnlyHint=True,
    destructiveHint=False,
    idempotentHint=True,
    openWorldHint=False,
)
ARTIFACT_CREATE = ToolAnnotations(
    readOnlyHint=False,
    destructiveHint=False,
    idempotentHint=False,
    openWorldHint=False,
)

MAX_TEXT_FILE_BYTES = 2 * 1024 * 1024
MAX_CONTEXT_BYTES = 8 * 1024 * 1024
MAX_OUTLINE_BYTES = 2 * 1024 * 1024
MAX_EXCEL_BYTES = 32 * 1024 * 1024
MAX_EXCEL_UNCOMPRESSED_BYTES = 256 * 1024 * 1024
MAX_EXCEL_ENTRIES = 10_000
MAX_EXCEL_CELLS = 10_000
MAX_DINGO_BYTES = 16 * 1024 * 1024
MAX_DINGO_RECORDS = 5_000
MAX_DINGO_CONTENT_CHARS = 20_000
MAX_DINGO_TOTAL_CHARS = 2_000_000
MAX_INLINE_RESULT_BYTES = 220_000
FIXED_ZIP_TIME = (1980, 1, 1, 0, 0, 0)
FIXED_CORE_TIMESTAMP = b"2000-01-01T00:00:00Z"
CORE_TIME_PATTERN = re.compile(
    rb"(<dcterms:(?:created|modified)\b[^>]*>).*?"
    rb"(</dcterms:(?:created|modified)>)",
    re.DOTALL,
)
SAFE_CONTENT_FIELD = re.compile(r"[A-Za-z_][A-Za-z0-9_]{0,63}")
SAFE_SHEET_NAME = re.compile(r"[^\\/*?:\[\]]{1,31}")
CODE_EXTENSIONS = {
    ".c", ".cc", ".cpp", ".cs", ".css", ".go", ".h", ".hpp",
    ".html", ".java", ".js", ".json", ".jsx", ".kt", ".kts",
    ".md", ".markdown", ".php", ".py", ".rb", ".rs", ".scss",
    ".sh", ".sql", ".swift", ".toml", ".ts", ".tsx", ".txt",
    ".vue", ".xml", ".yaml", ".yml",
}
SPECIAL_CHARACTER_PATTERNS = tuple(
    re.compile(value)
    for value in (
        r"u200e",
        r"&#247;|\? :",
        r"[�□]|\{\/U\}",
        r"U\+26[0-F][0-D]|U\+273[3-4]|U\+1F[3-6][0-4][0-F]|U\+1F6[8-F][0-F]",
        r"<\|.*?\|>",
    )
)


class _StrictModel(BaseModel):
    model_config = ConfigDict(extra="forbid", str_strip_whitespace=True)


def _freeze_strict_tool_contract(mcp: FastMCP) -> FastMCP:
    for tool in mcp._tool_manager._tools.values():
        argument_model = tool.fn_metadata.arg_model
        argument_model.model_config = ConfigDict(
            **dict(argument_model.model_config),
            extra="forbid",
        )
        argument_model.model_rebuild(force=True)
        tool.parameters = argument_model.model_json_schema(by_alias=True)
    return mcp


def _workspace_items(context: Any) -> tuple[tuple[str, str, Path], ...]:
    items = context.workspace_files()
    if len(items) > 5_000:
        raise ValueError("工作区文件数超过 5000 个上限。")
    return items


def _read_code_file(path: Path) -> str:
    if path.suffix.casefold() not in CODE_EXTENSIONS:
        raise ValueError("工作区包含不受支持的代码或文本扩展名。")
    data = path.read_bytes()
    if len(data) > MAX_TEXT_FILE_BYTES or b"\x00" in data:
        raise ValueError("代码文件超过 2 MiB 或包含二进制内容。")
    try:
        return data.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise ValueError("代码文件必须使用 UTF-8。") from exc


def _python_outline(text: str) -> list[str]:
    try:
        tree = ast.parse(text)
    except (SyntaxError, ValueError, MemoryError, RecursionError):
        return []
    values: list[str] = []
    for node in tree.body:
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
            values.append(f"{node.__class__.__name__} {node.name} (line {node.lineno})")
        elif isinstance(node, ast.ClassDef):
            values.append(f"class {node.name} (line {node.lineno})")
    return values


def _text_outline(path: str, text: str) -> list[str]:
    suffix = Path(path).suffix.casefold()
    if suffix == ".py":
        return _python_outline(text)
    if suffix in {".md", ".markdown"}:
        return [
            f"heading {line.strip()} (line {index})"
            for index, line in enumerate(text.splitlines(), 1)
            if re.fullmatch(r"#{1,6}\s+.{1,200}", line.strip())
        ]
    if suffix == ".json":
        try:
            value = json.loads(text)
        except (ValueError, RecursionError):
            return []
        if isinstance(value, dict):
            return [f"key {key}" for key in list(value)[:200] if isinstance(key, str)]
        if isinstance(value, list):
            return [f"array items={len(value)}"]
        return [f"scalar {type(value).__name__}"]
    declaration = re.compile(
        r"^\s*(?:export\s+)?(?:async\s+)?"
        r"(class|interface|enum|struct|type|function|def|func|fn)\s+"
        r"([A-Za-z_$][A-Za-z0-9_$]*)"
    )
    values = []
    for index, line in enumerate(text.splitlines()[:20_000], 1):
        match = declaration.match(line)
        if match:
            values.append(f"{match.group(1)} {match.group(2)} (line {index})")
    return values


def build_llm_context(context: Any) -> FastMCP:
    """Expose a sealed-workspace subset of llm-context 0.6.4."""

    mcp = FastMCP("ModelMirror llm-context")

    @mcp.tool(name="lc_preview", annotations=READ_ONLY)
    def lc_preview(max_files: int = 100) -> dict[str, Any]:
        """Preview supported files in the sealed workspace without accepting root_path or rules."""

        limit = max(1, min(int(max_files), 200))
        result: list[dict[str, Any]] = []
        total = 0
        for file_id, relative, path in _workspace_items(context):
            if path.suffix.casefold() not in CODE_EXTENSIONS:
                continue
            size = path.stat().st_size
            total += size
            result.append({"file_id": file_id, "path": relative, "size_bytes": size})
            if len(result) >= limit:
                break
        return {
            "rule": "modelmirror-sealed-workspace",
            "files": result,
            "file_count": len(result),
            "total_bytes": total,
            "truncated": len(result) >= limit,
        }

    @mcp.tool(name="lc_outlines", annotations=ARTIFACT_CREATE)
    def lc_outlines(
        artifact_name: ArtifactName = "workspace-outlines.md",
        max_files: int = 100,
        max_symbols_per_file: int = 200,
    ) -> dict[str, Any]:
        """Generate bounded code outlines from the sealed workspace as a registered artifact."""

        file_limit = max(1, min(int(max_files), 200))
        symbol_limit = max(1, min(int(max_symbols_per_file), 500))
        blocks = ["# Sealed workspace outlines", "", "Rule: `modelmirror-sealed-workspace`", ""]
        source_bytes = 0
        file_count = 0
        for _, relative, path in _workspace_items(context):
            if path.suffix.casefold() not in CODE_EXTENSIONS:
                continue
            text = _read_code_file(path)
            source_bytes += len(text.encode("utf-8"))
            if source_bytes > MAX_CONTEXT_BYTES:
                raise ValueError("代码上下文总量超过 8 MiB。")
            symbols = _text_outline(relative, text)[:symbol_limit]
            blocks.extend([f"## `{relative}`", ""])
            blocks.extend(f"- {value}" for value in symbols)
            if not symbols:
                blocks.append("- no supported declarations found")
            blocks.append("")
            file_count += 1
            if file_count >= file_limit:
                break
        if not file_count:
            raise ValueError("工作区没有可分析的 UTF-8 代码或文本文件。")
        data = ("\n".join(blocks).rstrip() + "\n").encode("utf-8")
        if len(data) > MAX_OUTLINE_BYTES:
            raise ValueError("Outline 产物超过 2 MiB。")
        target = context.artifact_path(artifact_name, ".md")
        target.write_bytes(data)
        payload = context.artifact_payload(target)
        payload.update({"file_count": file_count, "rule": "modelmirror-sealed-workspace"})
        return payload

    return _freeze_strict_tool_contract(mcp)


def _validate_xlsx_package(path: Path) -> None:
    if path.suffix.casefold() != ".xlsx":
        raise ValueError("Excel facade 仅接受 XLSX。")
    size = path.stat().st_size
    if size <= 0 or size > MAX_EXCEL_BYTES:
        raise ValueError("XLSX 必须非空且不超过 32 MiB。")
    try:
        with zipfile.ZipFile(path, "r") as archive:
            infos = archive.infolist()
            if len(infos) > MAX_EXCEL_ENTRIES:
                raise ValueError("XLSX ZIP 条目过多。")
            uncompressed = sum(info.file_size for info in infos)
            compressed = max(1, sum(info.compress_size for info in infos))
            if (
                uncompressed > MAX_EXCEL_UNCOMPRESSED_BYTES
                or uncompressed > compressed * 100
            ):
                raise ValueError("XLSX 解压规模或压缩比超过上限。")
            names = [info.filename.casefold() for info in infos]
            forbidden_parts = (
                "vbaproject",
                "externallinks/",
                "connections.xml",
                "querytables/",
                "embeddings/",
                "oleobjects/",
                "activex/",
            )
            if any(any(value in name for value in forbidden_parts) for name in names):
                raise ValueError("XLSX 包含宏、外部连接或嵌入对象。")
            for info in infos:
                lower = info.filename.casefold()
                if not (lower.endswith(".rels") or lower == "[content_types].xml"):
                    continue
                if info.file_size > 4 * 1024 * 1024:
                    raise ValueError("XLSX 关系清单过大。")
                content = archive.read(info).lower()
                if any(
                    token in content
                    for token in (
                        b'targetmode="external"',
                        b"targetmode='external'",
                        b"externallink",
                        b"vbaproject",
                        b"oleobject",
                        b"attachedtemplate",
                        b"macroenabled",
                    )
                ):
                    raise ValueError("XLSX 包含外部关系或活动内容。")
    except zipfile.BadZipFile as exc:
        raise ValueError("XLSX ZIP 结构无效。") from exc


def _excel_value(value: Any) -> str | int | float | bool | None:
    if value is None or isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, (datetime,)):
        return value.isoformat()
    return str(value)[:2_000]


def _range_bounds(start_cell: str, end_cell: str | None) -> tuple[int, int, int, int]:
    from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries

    try:
        start_row, start_column = coordinate_to_tuple(str(start_cell or ""))
        if end_cell:
            min_column, min_row, max_column, max_row = range_boundaries(
                f"{start_cell}:{end_cell}"
            )
        else:
            min_column, min_row = start_column, start_row
            max_column, max_row = start_column + 49, start_row + 199
    except (TypeError, ValueError) as exc:
        raise ValueError("Excel 单元格范围无效。") from exc
    if min_row < 1 or min_column < 1 or max_row < min_row or max_column < min_column:
        raise ValueError("Excel 单元格范围无效。")
    if (max_row - min_row + 1) * (max_column - min_column + 1) > MAX_EXCEL_CELLS:
        raise ValueError("Excel 读取范围最多 10000 个单元格。")
    return min_row, min_column, max_row, max_column


def _load_xlsx(path: Path, *, read_only: bool, data_only: bool) -> Any:
    from openpyxl import load_workbook

    _validate_xlsx_package(path)
    try:
        workbook = load_workbook(
            path,
            read_only=read_only,
            data_only=data_only,
            keep_links=False,
        )
    except Exception as exc:
        raise ValueError("XLSX 无法安全解析。") from exc
    if len(workbook.sheetnames) > 50:
        workbook.close()
        raise ValueError("XLSX 工作表数量超过 50。")
    return workbook


def _copy_deterministic_zip(source: Path, target: Path) -> None:
    with zipfile.ZipFile(source, "r") as archive, zipfile.ZipFile(
        target,
        "w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=9,
    ) as output:
        for info in sorted(archive.infolist(), key=lambda item: item.filename):
            if info.is_dir():
                continue
            data = archive.read(info)
            if info.filename.casefold() == "docprops/core.xml":
                data = CORE_TIME_PATTERN.sub(
                    lambda match: match.group(1)
                    + FIXED_CORE_TIMESTAMP
                    + match.group(2),
                    data,
                )
            clean = zipfile.ZipInfo(info.filename, FIXED_ZIP_TIME)
            clean.compress_type = zipfile.ZIP_DEFLATED
            clean.external_attr = 0o100644 << 16
            clean.create_system = 3
            output.writestr(clean, data)


def build_excel_analysis(context: Any) -> FastMCP:
    """Expose the reviewed file-copy subset of Excel MCP Server v0.1.8."""

    mcp = FastMCP("ModelMirror Excel MCP 0.1.8")

    @mcp.tool(name="get_workbook_metadata", annotations=READ_ONLY)
    def get_workbook_metadata(file_id: FileId, include_ranges: bool = False) -> dict[str, Any]:
        path = context.resolve_file(file_id)
        workbook = _load_xlsx(path, read_only=True, data_only=True)
        try:
            sheets = []
            for sheet in workbook.worksheets:
                item: dict[str, Any] = {
                    "name": sheet.title,
                    "max_row": min(int(sheet.max_row or 0), 100_000),
                    "max_column": min(int(sheet.max_column or 0), 500),
                }
                if include_ranges:
                    item["dimension"] = str(sheet.calculate_dimension())[:80]
                sheets.append(item)
            return {"sheet_count": len(sheets), "sheets": sheets}
        finally:
            workbook.close()

    @mcp.tool(name="read_data_from_excel", annotations=READ_ONLY)
    def read_data_from_excel(
        file_id: FileId,
        sheet_name: str,
        start_cell: str = "A1",
        end_cell: str | None = None,
    ) -> dict[str, Any]:
        if not SAFE_SHEET_NAME.fullmatch(str(sheet_name or "")):
            raise ValueError("工作表名称无效。")
        min_row, min_column, max_row, max_column = _range_bounds(start_cell, end_cell)
        path = context.resolve_file(file_id)
        workbook = _load_xlsx(path, read_only=True, data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError("工作表不存在。")
            sheet = workbook[sheet_name]
            rows = [
                [_excel_value(cell.value) for cell in row]
                for row in sheet.iter_rows(
                    min_row=min_row,
                    min_col=min_column,
                    max_row=max_row,
                    max_col=max_column,
                )
            ]
            payload = {
                "sheet_name": sheet_name,
                "start_cell": start_cell,
                "end_cell": end_cell,
                "rows": rows,
                "row_count": len(rows),
                "column_count": max((len(row) for row in rows), default=0),
            }
            if len(json.dumps(payload, ensure_ascii=False, default=str).encode("utf-8")) > MAX_INLINE_RESULT_BYTES:
                raise ValueError("Excel 范围结果超过 220000 字节，请缩小范围。")
            return payload
        finally:
            workbook.close()

    @mcp.tool(name="write_data_to_excel", annotations=ARTIFACT_CREATE)
    def write_data_to_excel(
        file_id: FileId,
        sheet_name: str,
        data: list[list[str | int | float | bool | None]],
        start_cell: str = "A1",
        artifact_name: ArtifactName = "updated.xlsx",
    ) -> dict[str, Any]:
        if not SAFE_SHEET_NAME.fullmatch(str(sheet_name or "")):
            raise ValueError("工作表名称无效。")
        if not data or len(data) > 1_000:
            raise ValueError("写入数据必须包含 1 到 1000 行。")
        width = max((len(row) for row in data), default=0)
        if width < 1 or width > 100 or sum(len(row) for row in data) > MAX_EXCEL_CELLS:
            raise ValueError("写入数据最多 100 列、10000 个单元格。")
        start_row, start_column, _, _ = _range_bounds(start_cell, start_cell)
        for row in data:
            for value in row:
                if isinstance(value, str) and (len(value) > 5_000 or value.startswith("=")):
                    raise ValueError("Excel 文本过长或包含公式。")
                if isinstance(value, float) and not math.isfinite(value):
                    raise ValueError("Excel 数值必须有限。")
        source = context.resolve_file(file_id)
        workbook = _load_xlsx(source, read_only=False, data_only=False)
        target = context.artifact_path(artifact_name, ".xlsx")
        raw = target.with_name(f".{target.stem}.raw.xlsx")
        if target.exists() or raw.exists():
            workbook.close()
            raise ValueError("同名 Excel 产物已经存在。")
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError("工作表不存在。")
            existing = 0
            for sheet in workbook.worksheets:
                existing += int(sheet.max_row or 0) * int(sheet.max_column or 0)
                if existing > 100_000:
                    raise ValueError("源工作簿已用单元格超过 100000。")
                for row in sheet.iter_rows():
                    if any(cell.data_type == "f" for cell in row):
                        raise ValueError("源工作簿包含公式，不能生成输出副本。")
            sheet = workbook[sheet_name]
            for row_offset, values in enumerate(data):
                for column_offset, value in enumerate(values):
                    sheet.cell(
                        row=start_row + row_offset,
                        column=start_column + column_offset,
                        value=value,
                    )
            workbook.properties.created = datetime(2000, 1, 1)
            workbook.properties.modified = datetime(2000, 1, 1)
            workbook.save(raw)
            _validate_xlsx_package(raw)
            _copy_deterministic_zip(raw, target)
            _validate_xlsx_package(target)
        except Exception:
            target.unlink(missing_ok=True)
            raise
        finally:
            workbook.close()
            raw.unlink(missing_ok=True)
        payload = context.artifact_payload(target)
        payload.update(
            {
                "sheet_name": sheet_name,
                "start_cell": start_cell,
                "written_cells": sum(len(row) for row in data),
            }
        )
        return payload

    return _freeze_strict_tool_contract(mcp)


def _dingo_records(path: Path, content_field: str) -> list[str]:
    if not SAFE_CONTENT_FIELD.fullmatch(content_field):
        raise ValueError("content_field 必须是简单字段名。")
    data = path.read_bytes()
    if not data or len(data) > MAX_DINGO_BYTES or b"\x00" in data:
        raise ValueError("Dingo 输入必须非空、非二进制且不超过 16 MiB。")
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise ValueError("Dingo 输入必须使用 UTF-8。") from exc
    suffix = path.suffix.casefold()
    raw_records: list[Any]
    if suffix == ".txt":
        raw_records = text.splitlines() or [""]
    elif suffix == ".jsonl":
        raw_records = []
        for line in text.splitlines():
            if line.strip():
                raw_records.append(json.loads(line))
    elif suffix == ".json":
        value = json.loads(text)
        if not isinstance(value, list):
            raise ValueError("JSON 输入顶层必须是数组。")
        raw_records = value
    elif suffix == ".csv":
        raw_records = list(csv.DictReader(io.StringIO(text)))
    else:
        raise ValueError("Dingo facade 仅接受 JSONL、JSON、CSV 或 TXT。")
    if not raw_records or len(raw_records) > MAX_DINGO_RECORDS:
        raise ValueError("Dingo 输入必须包含 1 到 5000 条记录。")
    records: list[str] = []
    total = 0
    for item in raw_records:
        if isinstance(item, str):
            content = item
        elif isinstance(item, dict):
            value = item.get(content_field)
            content = "" if value is None else value
        else:
            raise ValueError("Dingo 记录必须是字符串或对象。")
        if not isinstance(content, str):
            raise ValueError("Dingo 内容字段必须是字符串或 null。")
        if len(content) > MAX_DINGO_CONTENT_CHARS:
            raise ValueError("单条 Dingo 内容超过 20000 字符。")
        total += len(content)
        if total > MAX_DINGO_TOTAL_CHARS:
            raise ValueError("Dingo 内容总量超过 2000000 字符。")
        records.append(content)
    return records


def _dingo_issues(content: str) -> list[str]:
    issues = []
    if not content.strip():
        issues.append("RuleContentNull")
    if content and content[-1] == ":":
        issues.append("RuleColonEnd")
    matches = sum(len(pattern.findall(content)) for pattern in SPECIAL_CHARACTER_PATTERNS)
    if content and matches / len(content) >= 0.01:
        issues.append("RuleSpecialCharacter")
    return issues


def build_dingo_rules(context: Any) -> FastMCP:
    """Expose a fixed rule-only compatibility subset of Dingo v2.5.0."""

    mcp = FastMCP("ModelMirror Dingo Rules")

    @mcp.tool(name="list_dingo_components", annotations=READ_ONLY)
    def list_dingo_components(
        component_type: Literal["rule_groups"] = "rule_groups",
        include_details: bool = False,
    ) -> dict[str, Any]:
        group: dict[str, Any] = {"name": "modelmirror-basic"}
        if include_details:
            group.update(
                {
                    "rule_count": 3,
                    "rules": [
                        "RuleContentNull",
                        "RuleColonEnd",
                        "RuleSpecialCharacter",
                    ],
                }
            )
        return {"component_type": component_type, "rule_groups": [group]}

    @mcp.tool(name="run_dingo_evaluation", annotations=ARTIFACT_CREATE)
    def run_dingo_evaluation(
        file_id: FileId,
        content_field: str = "content",
        eval_group_name: Literal["modelmirror-basic"] = "modelmirror-basic",
        artifact_name: ArtifactName = "dingo-report.json",
    ) -> dict[str, Any]:
        source = context.resolve_file(file_id)
        records = _dingo_records(source, content_field)
        details = []
        issue_count = 0
        for index, content in enumerate(records, 1):
            issues = _dingo_issues(content)
            if issues:
                issue_count += 1
                details.append({"record": index, "rules": issues})
        report = {
            "evaluation_type": "rule",
            "eval_group_name": eval_group_name,
            "rules": ["RuleContentNull", "RuleColonEnd", "RuleSpecialCharacter"],
            "total": len(records),
            "num_bad": issue_count,
            "num_good": len(records) - issue_count,
            "score": round((len(records) - issue_count) * 100 / len(records), 6),
            "issues": details[:1_000],
            "issues_truncated": len(details) > 1_000,
        }
        data = (json.dumps(report, ensure_ascii=False, sort_keys=True, indent=2) + "\n").encode(
            "utf-8"
        )
        target = context.artifact_path(artifact_name, ".json")
        if target.exists():
            raise ValueError("同名 Dingo 产物已经存在。")
        target.write_bytes(data)
        payload = context.artifact_payload(target)
        payload.update(
            {
                "evaluation_type": "rule",
                "eval_group_name": eval_group_name,
                "total": len(records),
                "num_bad": issue_count,
                "score": report["score"],
            }
        )
        return payload

    return _freeze_strict_tool_contract(mcp)


WAVE18B_BUILDERS = {
    "cyberchitta-llm-context-py": build_llm_context,
    "haris-musa-excel-mcp-server": build_excel_analysis,
    "dataeval-dingo": build_dingo_rules,
}

WAVE18B_TOOL_NAMES = {
    "cyberchitta-llm-context-py": ("lc_preview", "lc_outlines"),
    "haris-musa-excel-mcp-server": (
        "get_workbook_metadata",
        "read_data_from_excel",
        "write_data_to_excel",
    ),
    "dataeval-dingo": ("list_dingo_components", "run_dingo_evaluation"),
}

WAVE18B_SCHEMA_SHA256 = {
    "cyberchitta-llm-context-py": (
        "f4978faaad49bc6d1a0ae9a3ba8da07dd404419e17049ff4266192014e42ebc7"
    ),
    "haris-musa-excel-mcp-server": (
        "81342cfe381afddab1f646ddab181b35d6eb767c4805f4191695f53f0be6f1f8"
    ),
    "dataeval-dingo": (
        "0a2f5e40c241efc77ba2b0728b550c904241b7e5ef2d080deca3e70ba9787b67"
    ),
}
