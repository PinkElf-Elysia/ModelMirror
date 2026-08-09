from __future__ import annotations

import io
import re
import zipfile
from datetime import date
from pathlib import Path

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook

from server.file_assets import document_parser as parser_module
from server.file_assets.api import router as file_asset_router
from server.file_assets.contracts import FileInputKind, FilePurpose
from server.file_assets.document_parser import (
    LocalDocumentParseError,
    parse_chat_document,
)
from server.file_assets.registry import get_file_format_registry
from server.file_assets.service import FileAssetService, get_file_asset_service
from server.file_assets.validation import FileUploadValidator, FileValidationError
from server.rag.document_parser import parse_document_structured
from server.rag.document_processor import StructuredDocumentProcessor


XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


@pytest.fixture(autouse=True)
def _enable_chat_files(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("CHAT_FILE_INPUT_ENABLED", "true")
    monkeypatch.setenv("FILE_ASSET_STORE_MODE", "shadow")


def _golden_workbook(path: Path) -> Path:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "概览"
    summary.append(["指标", "值", "日期"])
    summary.append(["缓存公式", "=1+2", date(2026, 8, 7)])
    summary.append(["无缓存公式", "=SUM(1, 4)", 5.5])
    summary.merge_cells("A4:B4")
    summary["A4"] = "合并标题"

    data = workbook.create_sheet("中文数据")
    data.append(["城市", "数量"])
    data.append(["上海", 42])

    hidden = workbook.create_sheet("内部计算")
    hidden["A1"] = "不应出现在预览中"
    hidden.sheet_state = "hidden"
    workbook.save(path)

    # openpyxl intentionally does not calculate formulas. Insert one genuine
    # cached value into the OOXML so data_only=True can prove the preference;
    # B3 remains uncached and must fall back to its inert formula text.
    _rewrite_zip_member(
        path,
        "xl/worksheets/sheet1.xml",
        lambda content: _replace_formula_cache(content, coordinate="B2", value="3"),
    )
    return path


def _replace_formula_cache(content: bytes, *, coordinate: str, value: str) -> bytes:
    empty_pattern = re.compile(
        rb'(<c\s+r="'
        + re.escape(coordinate.encode("ascii"))
        + rb'"[^>]*>\s*<f[^>]*>.*?</f>\s*)<v\s*/>',
        re.DOTALL,
    )
    updated, count = empty_pattern.subn(
        lambda match: match.group(1)
        + b"<v>"
        + value.encode("ascii")
        + b"</v>",
        content,
        count=1,
    )
    assert count == 1, f"formula cell {coordinate} was not found"
    return updated


def _rewrite_zip_member(
    path: Path,
    member_name: str,
    transform,
) -> None:
    output = io.BytesIO()
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        output, "w", zipfile.ZIP_DEFLATED
    ) as target:
        found = False
        for entry in source.infolist():
            content = source.read(entry.filename)
            if entry.filename == member_name:
                content = transform(content)
                found = True
            target.writestr(entry, content)
        assert found, f"missing OOXML member: {member_name}"
    path.write_bytes(output.getvalue())


def _add_zip_member(path: Path, member_name: str, content: bytes) -> bytes:
    output = io.BytesIO()
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        output, "w", zipfile.ZIP_DEFLATED
    ) as target:
        for entry in source.infolist():
            target.writestr(entry, source.read(entry.filename))
        target.writestr(member_name, content)
    return output.getvalue()


def _validate(
    content: bytes,
    *,
    purpose: FilePurpose,
) -> str:
    result = FileUploadValidator().validate_stream(
        io.BytesIO(content),
        purpose=purpose,
        input_kind=(
            FileInputKind.DATA_SOURCE
            if purpose == FilePurpose.DATAX
            else FileInputKind.DOCUMENT
        ),
        filename="workbook.xlsx",
        declared_media_type=XLSX_MEDIA_TYPE,
    )
    return result.format_id


def test_registry_exposes_xlsx_to_chat_rag_datax_and_workflow() -> None:
    registry = get_file_format_registry()
    assert registry.version == "modelmirror-file-formats-v4"

    for purpose in (FilePurpose.CHAT, FilePurpose.RAG):
        policy = next(
            item
            for item in registry.policies_for(purpose)
            if item.input_kind == FileInputKind.DOCUMENT
        )
        assert "xlsx" in policy.format_ids
        assert policy.max_bytes_per_file == 10 * 1024 * 1024

    agent_policy = next(
        item
        for item in registry.policies_for(FilePurpose.AGENT)
        if item.input_kind == FileInputKind.DOCUMENT
    )
    assert "xlsx" not in agent_policy.format_ids

    workflow_policy = next(
        item
        for item in registry.policies_for(FilePurpose.WORKFLOW)
        if item.input_kind == FileInputKind.DOCUMENT
    )
    assert "xlsx" in workflow_policy.format_ids
    assert workflow_policy.max_bytes_per_file == 10 * 1024 * 1024

    datax = next(
        item
        for item in registry.policies_for(FilePurpose.DATAX)
        if item.input_kind == FileInputKind.DATA_SOURCE
    )
    assert "xlsx" in datax.format_ids
    assert datax.max_bytes_per_file == 50 * 1024 * 1024


def test_xlsx_semantic_preview_preserves_sources_values_and_warnings(
    tmp_path: Path,
) -> None:
    path = _golden_workbook(tmp_path / "golden.xlsx")

    parsed = parse_document_structured(path, "golden.xlsx")

    assert parsed.format == "xlsx"
    assert {section.sheet for section in parsed.sections} == {"概览", "中文数据"}
    summary = next(section for section in parsed.sections if section.sheet == "概览")
    data = next(section for section in parsed.sections if section.sheet == "中文数据")
    assert summary.row_range == "A1:C4"
    assert "B2: 3" in summary.text
    assert "=1+2" not in summary.text
    assert "B3: =SUM(1, 4)" in summary.text
    assert "C2: 2026-08-07" in summary.text
    assert "C3: 5.5" in summary.text
    assert "A4: 合并标题" in summary.text
    assert "B4:" not in summary.text
    assert data.row_range == "A1:B2"
    assert "A2: 上海" in data.text
    assert all(section.sheet != "内部计算" for section in parsed.sections)
    assert any("隐藏工作表" in warning and "内部计算" in warning for warning in parsed.warnings)
    assert any("没有缓存结果" in warning and "未执行公式" in warning for warning in parsed.warnings)


def test_xlsx_loader_contract_is_explicitly_read_only_and_link_free(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = _golden_workbook(tmp_path / "contract.xlsx")
    calls: list[dict[str, object]] = []
    original = parser_module.load_workbook

    def recording_loader(*args, **kwargs):
        calls.append(dict(kwargs))
        return original(*args, **kwargs)

    monkeypatch.setattr(parser_module, "load_workbook", recording_loader)

    parse_chat_document(path, format_id="xlsx", title="contract.xlsx")

    assert calls == [
        {"read_only": True, "data_only": False, "keep_links": False},
        {"read_only": True, "data_only": True, "keep_links": False},
    ]


def test_chat_file_api_parses_xlsx_from_opaque_blob_key(tmp_path: Path) -> None:
    workbook_path = _golden_workbook(tmp_path / "api-workbook.xlsx")
    service = FileAssetService(storage_dir=tmp_path / "asset-store", mode="native")
    app = FastAPI()
    app.include_router(file_asset_router)
    app.dependency_overrides[get_file_asset_service] = lambda: service

    with TestClient(app) as client:
        upload = client.post(
            "/api/files",
            data={"purpose": "chat", "scope_id": "xlsx-blob-session"},
            files={
                "file": (
                    "api-workbook.xlsx",
                    workbook_path.read_bytes(),
                    XLSX_MEDIA_TYPE,
                )
            },
        )
        assert upload.status_code == 201, upload.text
        asset_id = upload.json()["asset_id"]
        stored = service.repository.get_asset("local", asset_id)
        assert stored is not None
        assert Path(stored.storage_key).suffix == ".blob"

        query = "?purpose=chat&scope_id=xlsx-blob-session"
        parsed = client.post(f"/api/files/{asset_id}/parse{query}")
        assert parsed.status_code == 200, parsed.text
        preview = client.get(f"/api/files/{asset_id}/preview{query}")
        assert preview.status_code == 200, preview.text
        sections = preview.json()["sections"]
        summary = next(item for item in sections if item["sheet"] == "概览")
        assert summary["row_range"] == "A1:C4"
        assert "B2: 3" in summary["text"]


def test_xlsx_rag_blocks_keep_sheet_and_cell_range_metadata(tmp_path: Path) -> None:
    path = _golden_workbook(tmp_path / "rag.xlsx")

    processed = StructuredDocumentProcessor().process(
        path,
        filename="rag.xlsx",
        source_id="xlsx_source",
    )

    assert processed.blocks
    summary = next(block for block in processed.blocks if block.metadata.get("sheet") == "概览")
    assert summary.kind == "table"
    assert summary.metadata["row_range"] == "A1:C4"
    assert "B2: 3" in summary.text
    assert any("隐藏工作表" in warning for warning in processed.warnings)


def test_xlsx_semantic_limits_are_stable_and_datax_is_not_narrowed(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    wide = Workbook()
    wide.active["C1"] = "third"
    wide_path = tmp_path / "wide.xlsx"
    wide.save(wide_path)

    # OOXML safety validation is shared and does not impose the Chat/RAG
    # semantic column limit on Data X's specialized importer.
    assert _validate(wide_path.read_bytes(), purpose=FilePurpose.DATAX) == "xlsx"
    monkeypatch.setattr(parser_module, "MAX_XLSX_COLUMNS", 2)
    with pytest.raises(LocalDocumentParseError) as captured:
        parse_chat_document(wide_path, format_id="xlsx", title="wide.xlsx")
    assert captured.value.error_code == "xlsx_column_limit_exceeded"

    cells = Workbook()
    cells.active.append(["one", "two", "three"])
    cells_path = tmp_path / "cells.xlsx"
    cells.save(cells_path)
    monkeypatch.setattr(parser_module, "MAX_XLSX_COLUMNS", 200)
    monkeypatch.setattr(parser_module, "MAX_XLSX_NONEMPTY_CELLS", 2)
    with pytest.raises(LocalDocumentParseError) as captured:
        parse_chat_document(cells_path, format_id="xlsx", title="cells.xlsx")
    assert captured.value.error_code == "xlsx_cell_limit_exceeded"


def test_xlsx_visible_sheet_limit_ignores_hidden_sheets(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active["A1"] = "visible"
    for index in range(50):
        sheet = workbook.create_sheet(f"hidden-{index}")
        sheet["A1"] = "ignored"
        sheet.sheet_state = "hidden"
    path = tmp_path / "hidden.xlsx"
    workbook.save(path)
    assert parse_chat_document(path, format_id="xlsx", title="hidden.xlsx").sections

    for index in range(50):
        workbook.create_sheet(f"visible-{index}")["A1"] = index
    path = tmp_path / "too-many-sheets.xlsx"
    workbook.save(path)
    with pytest.raises(LocalDocumentParseError) as captured:
        parse_chat_document(path, format_id="xlsx", title=path.name)
    assert captured.value.error_code == "xlsx_sheet_limit_exceeded"


@pytest.mark.parametrize("purpose", (FilePurpose.CHAT, FilePurpose.RAG))
def test_xlsx_ooxml_preflight_rejects_damage_macros_and_external_links(
    tmp_path: Path,
    purpose: FilePurpose,
) -> None:
    path = _golden_workbook(tmp_path / f"safe-{purpose.value}.xlsx")
    with pytest.raises(FileValidationError) as captured:
        _validate(b"PK\x03\x04broken", purpose=purpose)
    assert captured.value.error_code == "invalid_xlsx"

    macro = _add_zip_member(path, "xl/vbaProject.bin", b"not-executed")
    with pytest.raises(FileValidationError) as captured:
        _validate(macro, purpose=purpose)
    assert captured.value.error_code == "unsupported_xlsx_feature"

    external_path = tmp_path / f"external-{purpose.value}.xlsx"
    external_path.write_bytes(path.read_bytes())

    def add_external_relationship(content: bytes) -> bytes:
        relationship = (
            b'<Relationship Id="external" Type="urn:modelmirror:test" '
            b'Target="https://example.invalid/book.xlsx" TargetMode="External"/>'
        )
        return content.replace(b"</Relationships>", relationship + b"</Relationships>")

    _rewrite_zip_member(
        external_path,
        "xl/_rels/workbook.xml.rels",
        add_external_relationship,
    )
    with pytest.raises(FileValidationError) as captured:
        _validate(external_path.read_bytes(), purpose=purpose)
    assert captured.value.error_code == "unsupported_xlsx_feature"
