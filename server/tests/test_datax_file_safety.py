from __future__ import annotations

import asyncio
import io
import json
import zipfile
from pathlib import Path
from typing import Any

import duckdb
import pytest
import starlette.formparsers
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook

import server.datax.service as datax_service_module
import server.datax.api as datax_api
from server.datax.api import configure_datax, router
from server.datax.service import DataXService, _xlsx_rows_with_formula_fallback
from server.datax.store import DataXStore, DataXUploadValidationError


XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _service(tmp_path: Path) -> DataXService:
    return DataXService(DataXStore(tmp_path / "datax"))


def _xlsx_bytes(*, empty: bool = False) -> bytes:
    workbook = Workbook()
    if not empty:
        workbook.active.append(["name", "amount"])
        workbook.active.append(["Alpha", 7])
    payload = io.BytesIO()
    workbook.save(payload)
    workbook.close()
    return payload.getvalue()


def _with_zip_member(payload: bytes, name: str, content: bytes) -> bytes:
    source = zipfile.ZipFile(io.BytesIO(payload), "r")
    target = io.BytesIO()
    with source, zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as archive:
        for entry in source.infolist():
            archive.writestr(entry, source.read(entry.filename))
        archive.writestr(name, content)
    return target.getvalue()


def _client(service: DataXService) -> TestClient:
    configure_datax(service)
    app = FastAPI()
    app.include_router(router)
    return TestClient(app)


class _UnexpectedDataXService:
    def __init__(self) -> None:
        self.create_calls = 0

    def create_import_job(self, *_args, **_kwargs):
        self.create_calls += 1
        raise AssertionError("Data X service must not run for an oversized request")


def test_datax_api_rejects_mime_and_unsafe_xlsx_before_persistence(
    tmp_path: Path,
) -> None:
    service = _service(tmp_path)
    project = service.create_project(name="Safety")
    client = _client(service)
    endpoint = f"/api/datax/projects/{project.project_id}/sources"
    valid = _xlsx_bytes()

    mismatch = client.post(
        endpoint,
        files={"file": ("facts.xlsx", valid, "text/plain")},
    )
    assert mismatch.status_code == 415
    assert mismatch.headers["x-modelmirror-error-code"] == "mime_type_mismatch"

    macro = _with_zip_member(valid, "xl/vbaProject.bin", b"not-a-real-macro")
    unsafe = client.post(
        endpoint,
        files={"file": ("facts.xlsx", macro, XLSX_MEDIA_TYPE)},
    )
    assert unsafe.status_code == 422
    assert unsafe.headers["x-modelmirror-error-code"] == "unsupported_xlsx_feature"

    fake_parquet = client.post(
        endpoint,
        files={
            "file": (
                "facts.parquet",
                b"PAR1not-real-metadataPAR1",
                "application/vnd.apache.parquet",
            )
        },
    )
    assert fake_parquet.status_code == 422
    assert not service.list_sources(project.project_id)
    assert not service.list_import_jobs(project.project_id)
    assert not list(service.store.sources_dir.rglob("*.*"))


def test_declared_datax_request_limit_rejects_before_multipart_or_service() -> None:
    service = _UnexpectedDataXService()
    configure_datax(service)  # type: ignore[arg-type]
    app = FastAPI()
    app.include_router(router)

    with TestClient(app) as client:
        response = client.post(
            "/api/datax/projects/project_1/sources",
            content=b"not-parsed",
            headers={
                "content-type": "multipart/form-data; boundary=unused",
                "content-length": str(datax_api.MAX_DATAX_UPLOAD_REQUEST_BYTES + 1),
            },
        )

    assert response.status_code == 413
    assert response.headers["x-modelmirror-error-code"] == "file_request_too_large"
    assert service.create_calls == 0


def test_chunked_datax_request_stops_and_closes_spool(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    boundary = b"modelmirror-datax-boundary"
    prefix = (
        b"--"
        + boundary
        + b'\r\nContent-Disposition: form-data; name="file"; filename="facts.csv"\r\n'
        b"Content-Type: text/csv\r\n\r\n"
    )
    first_file_chunk = b"a" * 512
    rejected_file_chunk = b"b" * 256
    first_request_chunk = prefix + first_file_chunk
    request_limit = len(first_request_chunk) + 32
    full_file_bytes = len(first_file_chunk) + len(rejected_file_chunk)
    request_chunks = [
        first_request_chunk,
        rejected_file_chunk,
        b"\r\n--" + boundary + b"--\r\n",
    ]

    real_spooled_file = starlette.formparsers.SpooledTemporaryFile
    tracked_files: list[Any] = []
    written_bytes = 0

    class TrackingSpooledFile:
        def __init__(self, *args, **kwargs) -> None:
            self._inner = real_spooled_file(*args, **kwargs)
            tracked_files.append(self)

        def write(self, content: bytes) -> int:
            nonlocal written_bytes
            written_bytes += len(content)
            return self._inner.write(content)

        def __getattr__(self, name: str):
            return getattr(self._inner, name)

    monkeypatch.setattr(
        starlette.formparsers, "SpooledTemporaryFile", TrackingSpooledFile
    )
    monkeypatch.setattr(
        datax_api, "MAX_DATAX_UPLOAD_REQUEST_BYTES", request_limit
    )

    service = _UnexpectedDataXService()
    configure_datax(service)  # type: ignore[arg-type]
    app = FastAPI()
    app.include_router(router)
    sent: list[dict[str, Any]] = []
    incoming = [
        {
            "type": "http.request",
            "body": chunk,
            "more_body": index < len(request_chunks) - 1,
        }
        for index, chunk in enumerate(request_chunks)
    ]

    async def receive() -> dict[str, Any]:
        if incoming:
            return incoming.pop(0)
        return {"type": "http.disconnect"}

    async def send(message: dict[str, Any]) -> None:
        sent.append(message)

    scope: dict[str, Any] = {
        "type": "http",
        "asgi": {"version": "3.0", "spec_version": "2.4"},
        "http_version": "1.1",
        "method": "POST",
        "scheme": "http",
        "path": "/api/datax/projects/project_1/sources",
        "raw_path": b"/api/datax/projects/project_1/sources",
        "query_string": b"",
        "root_path": "",
        "headers": [
            (b"host", b"testserver"),
            (b"content-type", b"multipart/form-data; boundary=" + boundary),
        ],
        "client": ("testclient", 50000),
        "server": ("testserver", 80),
    }
    asyncio.run(app(scope, receive, send))

    response_start = next(
        message for message in sent if message["type"] == "http.response.start"
    )
    response_body = b"".join(
        message.get("body", b"")
        for message in sent
        if message["type"] == "http.response.body"
    )
    assert response_start["status"] == 413
    assert json.loads(response_body)["detail"].startswith("上传请求超过")
    assert service.create_calls == 0
    assert len(incoming) == 1
    assert 0 < written_bytes < full_file_bytes
    assert tracked_files and all(item.closed for item in tracked_files)


def test_failed_import_keeps_task_metadata_and_removes_source_content(
    tmp_path: Path,
) -> None:
    service = _service(tmp_path)
    project = service.create_project(name="Failed import")
    payload = _xlsx_bytes(empty=True)
    job = service.create_import_job(
        project.project_id,
        file_name="empty.xlsx",
        content=payload,
        media_type=XLSX_MEDIA_TYPE,
    )
    source = service.list_sources(project.project_id)[0]
    source_path = service.store.source_file_path(
        project.project_id, source.content_sha256, ".xlsx"
    )
    assert source_path.exists()

    failed = service.run_import_job(job.job_id)
    assert failed.status == "failed"
    assert failed.completed_at is not None
    assert not source_path.exists()

    restarted = DataXService(DataXStore(tmp_path / "datax"))
    persisted = restarted.list_import_jobs(project.project_id)
    assert [item.job_id for item in persisted] == [job.job_id]
    assert persisted[0].status == "failed"
    assert restarted.list_sources(project.project_id)[0].status == "failed"
    response = _client(restarted).get(
        f"/api/datax/projects/{project.project_id}/import-jobs"
    )
    assert response.status_code == 200
    assert response.json()["items"][0]["status"] == "failed"


def test_xlsx_uses_first_visible_sheet_and_preserves_uncached_formula_text(
    tmp_path: Path,
) -> None:
    service = _service(tmp_path)
    project = service.create_project(name="Workbook semantics")
    workbook = Workbook()
    ignored = workbook.active
    ignored.title = "其他数据"
    ignored.append(["ignored"])
    visible = workbook.create_sheet("数据")
    visible.append(["name", "calculated"])
    visible.append(["甲", "=1+6"])
    workbook.active = workbook.sheetnames.index("数据")
    hidden = workbook.create_sheet("隐藏")
    hidden.append(["secret"])
    hidden.sheet_state = "hidden"
    payload = io.BytesIO()
    workbook.save(payload)
    workbook.close()

    job = service.import_source(
        project.project_id,
        file_name="workbook.xlsx",
        content=payload.getvalue(),
        media_type=XLSX_MEDIA_TYPE,
    )
    assert job.status == "ready"
    source = service.list_sources(project.project_id)[0]
    metadata = source.profile["source"]
    assert metadata["selected_sheet"] == "数据"
    assert metadata["hidden_sheets_ignored"] == ["隐藏"]
    assert metadata["visible_sheets_ignored"] == ["其他数据"]
    assert metadata["formula_text_fallback_count"] == 1
    assert metadata["formula_execution"] is False
    assert metadata["external_links"] is False
    assert "formula_cache_missing_preserved_as_text" in metadata["warnings"]

    connection = duckdb.connect(
        str(service.store.project_db_path(project.project_id)), read_only=True
    )
    try:
        row = connection.execute(
            f'SELECT "name", "calculated" FROM "{source.table_name}"'
        ).fetchone()
    finally:
        connection.close()
    assert row == ("甲", "=1+6")


def test_formula_row_merge_prefers_cache_and_reports_fallback() -> None:
    class Sheet:
        def __init__(self, rows):
            self.rows = rows

        def iter_rows(self, *, values_only: bool):
            assert values_only is True
            return iter(self.rows)

    stats = {"cached_formula_count": 0, "formula_text_fallback_count": 0}
    rows = list(
        _xlsx_rows_with_formula_fallback(
            Sheet([(7, None)]),
            Sheet([("=1+6", "=2+3")]),
            stats=stats,
        )
    )
    assert rows == [(7, "=2+3")]
    assert stats == {"cached_formula_count": 1, "formula_text_fallback_count": 1}


def test_xlsx_uncached_formula_after_type_sample_is_preserved_as_text(
    tmp_path: Path,
) -> None:
    service = _service(tmp_path)
    project = service.create_project(name="Late formula")
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["value"])
    for value in range(2000):
        sheet.append([value])
    sheet.append(["=1+6"])
    payload = io.BytesIO()
    workbook.save(payload)
    workbook.close()

    job = service.import_source(
        project.project_id,
        file_name="late-formula.xlsx",
        content=payload.getvalue(),
        media_type=XLSX_MEDIA_TYPE,
    )

    assert job.status == "ready"
    source = service.list_sources(project.project_id)[0]
    assert source.row_count == 2001
    assert source.profile["source"]["formula_text_fallback_count"] == 1
    assert source.profile["columns"][0]["data_type"] == "VARCHAR"
    connection = duckdb.connect(
        str(service.store.project_db_path(project.project_id)), read_only=True
    )
    try:
        assert connection.execute(
            f'SELECT COUNT(*) FROM "{source.table_name}" WHERE "value" = ?',
            ["=1+6"],
        ).fetchone()[0] == 1
    finally:
        connection.close()


def test_datax_keeps_its_row_boundary_and_rejects_overflow_before_import(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service = _service(tmp_path)
    project = service.create_project(name="Boundaries")

    # This exceeds the Chat/RAG semantic preview's 100k-row boundary but is
    # intentionally valid for Data X's one-million-row analysis contract.
    content = b"value\n" + (b"1\n" * 100_001)
    ready = service.import_source(
        project.project_id,
        file_name="large.csv",
        content=content,
        media_type="text/csv",
    )
    assert ready.status == "ready"
    assert service.list_sources(project.project_id)[0].row_count == 100_001

    monkeypatch.setattr(datax_service_module, "MAX_SOURCE_ROWS", 1)
    with pytest.raises(DataXUploadValidationError) as captured:
        service.create_import_job(
            project.project_id,
            file_name="too-many.csv",
            content=b"value\n2\n3\n",
            media_type="text/csv",
        )
    assert captured.value.error_code == "csv_row_limit_exceeded"
    assert len(service.list_sources(project.project_id)) == 1
    assert len(service.list_import_jobs(project.project_id)) == 1


def test_datax_size_rejection_remains_a_validation_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service = _service(tmp_path)
    project = service.create_project(name="Size")
    monkeypatch.setattr(datax_service_module, "MAX_SOURCE_BYTES", 2)
    with pytest.raises(DataXUploadValidationError) as captured:
        service.create_import_job(
            project.project_id,
            file_name="small.csv",
            content=b"a\n1",
            media_type="text/csv",
        )
    assert captured.value.status_code == 413
    assert captured.value.error_code == "file_too_large"


def test_datax_csv_limits_are_rejected_before_source_persistence(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(datax_service_module, "MAX_SOURCE_ROWS", 1)
    monkeypatch.setattr(datax_service_module, "MAX_SOURCE_COLUMNS", 2)
    service = _service(tmp_path)
    project = service.create_project(name="CSV preflight")

    with pytest.raises(DataXUploadValidationError) as rows:
        service.create_import_job(
            project.project_id,
            file_name="rows.csv",
            content=b"a\n1\n2\n",
            media_type="text/csv",
        )
    assert rows.value.error_code == "csv_row_limit_exceeded"

    with pytest.raises(DataXUploadValidationError) as columns:
        service.create_import_job(
            project.project_id,
            file_name="columns.csv",
            content=b"a,b,c\n1,2,3\n",
            media_type="text/csv",
        )
    assert columns.value.error_code == "csv_column_limit_exceeded"
    assert service.list_sources(project.project_id) == []
    assert service.list_import_jobs(project.project_id) == []
    assert not list(service.store.sources_dir.rglob("*.*"))


def test_datax_parquet_limits_are_rejected_from_metadata_before_persistence(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(datax_service_module, "MAX_SOURCE_ROWS", 1)
    monkeypatch.setattr(datax_service_module, "MAX_SOURCE_COLUMNS", 1)
    service = _service(tmp_path)
    project = service.create_project(name="Parquet preflight")

    path = tmp_path / "rows.parquet"
    connection = duckdb.connect(database=":memory:")
    try:
        connection.execute(
            "COPY (SELECT range::INTEGER AS a FROM range(2)) TO ? (FORMAT PARQUET)",
            [str(path)],
        )
    finally:
        connection.close()
    with pytest.raises(DataXUploadValidationError) as rows:
        service.create_import_job(
            project.project_id,
            file_name="rows.parquet",
            content=path.read_bytes(),
            media_type="application/vnd.apache.parquet",
        )
    assert rows.value.error_code == "parquet_row_limit_exceeded"

    connection = duckdb.connect(database=":memory:")
    try:
        connection.execute(
            "COPY (SELECT 1::INTEGER AS a, 2::INTEGER AS b) TO ? (FORMAT PARQUET)",
            [str(path)],
        )
    finally:
        connection.close()
    with pytest.raises(DataXUploadValidationError) as columns:
        service.create_import_job(
            project.project_id,
            file_name="columns.parquet",
            content=path.read_bytes(),
            media_type="application/vnd.apache.parquet",
        )
    assert columns.value.error_code == "parquet_column_limit_exceeded"
    assert service.list_sources(project.project_id) == []
    assert service.list_import_jobs(project.project_id) == []


def test_datax_source_delete_unlinks_symlink_without_following_target(
    tmp_path: Path,
) -> None:
    store = DataXStore(tmp_path / "datax")
    target = store.source_file_path("project_b", "b" * 64, ".csv")
    target.write_bytes(b"keep")
    link = store.source_file_path("project_a", "a" * 64, ".csv")
    try:
        link.symlink_to(target)
    except (NotImplementedError, OSError):
        pytest.skip("symlinks are unavailable in this test environment")

    store.delete_source_file(link)

    assert target.read_bytes() == b"keep"
    assert not link.is_symlink()


def test_datax_metadata_save_failure_rolls_back_memory_and_blob(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service = _service(tmp_path)
    project = service.create_project(name="Atomic metadata")

    def fail_save() -> None:
        raise OSError("disk full")

    monkeypatch.setattr(service.store, "_save", fail_save)
    with pytest.raises(OSError, match="disk full"):
        service.create_import_job(
            project.project_id,
            file_name="facts.csv",
            content=b"a\n1\n",
            media_type="text/csv",
        )

    assert service.list_sources(project.project_id) == []
    assert service.list_import_jobs(project.project_id) == []
    assert not list(service.store.sources_dir.rglob("*.*"))


def test_broken_xlsx_is_rejected_before_persistence(tmp_path: Path) -> None:
    service = _service(tmp_path)
    project = service.create_project(name="Broken workbook")
    valid = _xlsx_bytes()
    source = zipfile.ZipFile(io.BytesIO(valid), "r")
    target = io.BytesIO()
    with source, zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as archive:
        for entry in source.infolist():
            if entry.filename != "xl/worksheets/sheet1.xml":
                archive.writestr(entry, source.read(entry.filename))

    with pytest.raises(DataXUploadValidationError) as captured:
        service.create_import_job(
            project.project_id,
            file_name="broken.xlsx",
            content=target.getvalue(),
            media_type=XLSX_MEDIA_TYPE,
        )
    assert captured.value.error_code == "invalid_xlsx"
    assert service.list_sources(project.project_id) == []
    assert service.list_import_jobs(project.project_id) == []


def test_failed_import_persists_only_stable_masked_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service = _service(tmp_path)
    project = service.create_project(name="Masked failure")
    job = service.create_import_job(
        project.project_id,
        file_name="facts.csv",
        content=b"a\n1\n",
        media_type="text/csv",
    )

    def fail_import(_source):
        raise RuntimeError("TOP_SECRET from /app/datax/storage/project/source.csv")

    monkeypatch.setattr(service, "_load_snapshot", fail_import)
    failed = service.run_import_job(job.job_id)

    assert failed.status == "failed"
    assert failed.error == "数据源解析失败，请检查文件结构、行列限制或重新导出后再试。"
    assert "TOP_SECRET" not in failed.error
    assert "/app/datax" not in failed.error
