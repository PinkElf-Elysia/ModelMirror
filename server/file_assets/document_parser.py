from __future__ import annotations

import csv
import io
import multiprocessing
import os
import re
import signal
from collections.abc import Callable
from datetime import date, datetime, time, timedelta
from html.parser import HTMLParser
from itertools import zip_longest
from multiprocessing.connection import Connection
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import BaseModel, ConfigDict, Field


MAX_EXTRACTED_CHARACTERS = 500_000
MAX_SECTION_CHARACTERS = 20_000
MAX_PDF_PAGE_CHARACTERS = 100_000
PDF_PARSE_TIMEOUT_SECONDS = 10.0
PDF_PARSE_WORKER_MEMORY_BYTES = 512 * 1024 * 1024
MAX_XLSX_VISIBLE_SHEETS = 50
MAX_XLSX_NONEMPTY_CELLS = 100_000
MAX_XLSX_COLUMNS = 200
# A sparse cell at an extreme row would make openpyxl materialize every missing
# read-only row. The semantic preview budget therefore also bounds the row span.
MAX_XLSX_ROW_SPAN = MAX_XLSX_NONEMPTY_CELLS

PdfWorkerTarget = Callable[[str, Connection, int, int, float], None]


class LocalDocumentParseError(ValueError):
    """Stable, user-facing local parsing failure."""

    def __init__(
        self,
        error_code: str,
        message: str,
        *,
        status_code: int = 422,
    ) -> None:
        super().__init__(message)
        self.error_code = error_code
        self.message = message
        self.status_code = status_code


class ParsedSection(BaseModel):
    model_config = ConfigDict(frozen=True)

    text: str = Field(min_length=1)
    page: int | None = Field(default=None, ge=1)
    slide: int | None = Field(default=None, ge=1)
    sheet: str | None = None
    line_range: str | None = None
    row_range: str | None = None
    heading_path: tuple[str, ...] | None = None
    time_range: str | None = None


class ParsedDocument(BaseModel):
    model_config = ConfigDict(frozen=True)

    format: str
    title: str | None = None
    sections: tuple[ParsedSection, ...]
    warnings: tuple[str, ...] = ()
    extracted_chars: int = Field(ge=0)
    truncated: bool = False


class ParsedDocumentPreview(ParsedDocument):
    asset_id: str
    artifact_id: str
    artifact_expires_at: str


def parse_chat_document(
    path: Path,
    *,
    format_id: str,
    title: str | None,
    office_parser: Any | None = None,
) -> ParsedDocument:
    """Parse a registry-backed local document without calling a provider."""

    if format_id in {
        "plain_text",
        "markdown",
        "json",
        "jsonl",
        "yaml",
        "xml",
        "source_code",
        "configuration",
        "log",
    }:
        return _parse_utf8_text(path, format_id=format_id, title=title)
    if format_id in {"csv", "tsv"}:
        return _parse_delimited_text(
            path,
            format_id=format_id,
            title=title,
            delimiter="," if format_id == "csv" else "\t",
        )
    if format_id == "html":
        return _parse_html_text(path, title=title)
    if format_id in {"srt", "vtt"}:
        return _parse_subtitle_text(path, format_id=format_id, title=title)
    if format_id == "xlsx":
        return _parse_xlsx_workbook(path, title=title)
    if format_id in {"docx", "pptx"}:
        # Office parsing is deliberately outside the API process.  Importing
        # the bridge lazily also keeps python-docx/python-pptx out of the main
        # runtime and makes a sidecar outage fail closed rather than falling
        # back to an in-process parser.
        from .office_sidecar import OfficeSidecarError, OfficeSidecarParser

        parser = office_parser or OfficeSidecarParser()
        try:
            return parser.parse(path, format_id=format_id, title=title)
        except OfficeSidecarError as exc:
            raise LocalDocumentParseError(
                exc.error_code,
                exc.message,
                status_code=exc.status_code,
            ) from exc
    if format_id == "pdf":
        return _parse_text_pdf(path, title=title)
    raise LocalDocumentParseError(
        "file_parse_not_supported",
        "当前文件格式尚未接入本地解析器。",
    )


def _parse_utf8_text(
    path: Path, *, format_id: str, title: str | None
) -> ParsedDocument:
    sections: list[ParsedSection] = []
    warnings: list[str] = []
    buffer: list[str] = []
    buffer_chars = 0
    first_line = 1
    last_line = 0
    retained_chars = 0
    extracted_chars = 0
    truncated = False

    try:
        with path.open("r", encoding="utf-8-sig", newline="") as source:
            for line_number, line in enumerate(source, start=1):
                extracted_chars += len(line)
                last_line = line_number
                if retained_chars >= MAX_EXTRACTED_CHARACTERS:
                    truncated = True
                    continue
                retained = line[: MAX_EXTRACTED_CHARACTERS - retained_chars]
                if len(retained) < len(line):
                    truncated = True
                retained_chars += len(retained)
                buffer.append(retained)
                buffer_chars += len(retained)
                if buffer_chars >= MAX_SECTION_CHARACTERS:
                    _append_text_section(
                        sections,
                        buffer,
                        first_line,
                        line_number,
                        preserve_layout=format_id
                        in {"json", "jsonl", "yaml", "xml", "source_code", "configuration", "log"},
                    )
                    buffer = []
                    buffer_chars = 0
                    first_line = line_number + 1
    except (OSError, UnicodeError) as exc:
        raise LocalDocumentParseError(
            "file_parse_failed",
            "文本内容无法安全读取，请确认文件未损坏并使用 UTF-8 编码。",
        ) from exc

    if buffer:
        _append_text_section(
            sections,
            buffer,
            first_line,
            last_line,
            preserve_layout=format_id
            in {"json", "jsonl", "yaml", "xml", "source_code", "configuration", "log"},
        )
    if not sections:
        raise LocalDocumentParseError(
            "file_has_no_readable_text",
            "文件中没有可读取的文字内容，请检查后重新上传。",
        )
    if truncated:
        warnings.append("内容超过 500,000 字符，预览已安全截断；发送时将按模型上下文继续裁剪。")
    return ParsedDocument(
        format=format_id,
        title=title or None,
        sections=tuple(sections),
        warnings=tuple(warnings),
        extracted_chars=extracted_chars,
        truncated=truncated,
    )


def _append_text_section(
    sections: list[ParsedSection],
    lines: list[str],
    first_line: int,
    last_line: int,
    *,
    preserve_layout: bool = False,
) -> None:
    raw_text = "".join(lines)
    text = raw_text if preserve_layout else raw_text.strip()
    if text.strip():
        sections.append(
            ParsedSection(
                text=text,
                line_range=f"{first_line}-{max(first_line, last_line)}",
            )
        )


def _read_utf8_text(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8-sig")
    except (OSError, UnicodeError) as exc:
        raise LocalDocumentParseError(
            "file_parse_failed",
            "文本内容无法安全读取，请确认文件未损坏并使用 UTF-8 编码。",
        ) from exc


def _parse_delimited_text(
    path: Path,
    *,
    format_id: str,
    title: str | None,
    delimiter: str,
) -> ParsedDocument:
    source_text = _read_utf8_text(path)
    sections: list[ParsedSection] = []
    warnings: list[str] = []
    retained = 0
    truncated = False
    header: list[str] | None = None
    section_lines: list[str] = []
    section_first_row = 1
    section_last_row = 0

    try:
        rows = csv.reader(io.StringIO(source_text, newline=""), delimiter=delimiter, strict=True)
        for row_number, row in enumerate(rows, start=1):
            if header is None:
                header = row
            rendered = _render_delimited_row(row, row_number=row_number)
            remaining = MAX_EXTRACTED_CHARACTERS - retained
            if remaining <= 0:
                truncated = True
                continue
            kept = rendered[:remaining]
            if len(kept) < len(rendered):
                truncated = True
            retained += len(kept)
            section_lines.append(kept)
            section_last_row = row_number
            if sum(len(line) for line in section_lines) >= MAX_SECTION_CHARACTERS:
                sections.append(
                    ParsedSection(
                        text="".join(section_lines).rstrip("\n"),
                        row_range=f"{section_first_row}-{section_last_row}",
                    )
                )
                section_lines = []
                section_first_row = row_number + 1
    except csv.Error as exc:
        raise LocalDocumentParseError(
            "file_parse_failed",
            f"{format_id.upper()} 结构无效，请检查引号、分隔符和换行后重试。",
        ) from exc

    if section_lines:
        sections.append(
            ParsedSection(
                text="".join(section_lines).rstrip("\n"),
                row_range=f"{section_first_row}-{max(section_first_row, section_last_row)}",
            )
        )
    if not sections:
        raise LocalDocumentParseError(
            "file_has_no_readable_text",
            f"{format_id.upper()} 中没有可读取的数据。",
        )
    if truncated:
        warnings.append("内容超过 500,000 字符，预览已安全截断；发送时将按模型上下文继续裁剪。")
    return ParsedDocument(
        format=format_id,
        title=title or None,
        sections=tuple(sections),
        warnings=tuple(warnings),
        extracted_chars=len(source_text),
        truncated=truncated,
    )


def _render_delimited_row(row: list[str], *, row_number: int) -> str:
    # Cells are always inert text. A leading spreadsheet formula marker is not
    # evaluated or rewritten; Markdown escaping only prevents layout breakage.
    values = [str(row_number), *row]
    escaped = [value.replace("\\", "\\\\").replace("|", "\\|").replace("\r", " ").replace("\n", " ") for value in values]
    return "| " + " | ".join(escaped) + " |\n"


def _parse_xlsx_workbook(path: Path, *, title: str | None) -> ParsedDocument:
    """Extract a bounded semantic preview without evaluating formulas or links."""

    formula_book = None
    cached_book = None
    try:
        # Managed blobs deliberately use an opaque `.blob` suffix. OpenPyXL
        # validates path extensions before opening them, so feed the already
        # validated OOXML bytes through independent seekable streams instead.
        workbook_bytes = path.read_bytes()
        # Both views are intentionally read-only and explicitly disable external
        # link preservation. data_only selects cached results; the formula view
        # is retained solely to recover formula text when a cache is absent.
        formula_book = load_workbook(
            io.BytesIO(workbook_bytes),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
        cached_book = load_workbook(
            io.BytesIO(workbook_bytes),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        visible_sheets = [
            worksheet
            for worksheet in formula_book.worksheets
            if worksheet.sheet_state == "visible"
        ]
        hidden_sheets = [
            worksheet.title
            for worksheet in formula_book.worksheets
            if worksheet.sheet_state != "visible"
        ]
        if len(visible_sheets) > MAX_XLSX_VISIBLE_SHEETS:
            raise LocalDocumentParseError(
                "xlsx_sheet_limit_exceeded",
                "XLSX 超过 50 个可见工作表，请拆分工作簿，或改用 Data X 分析。",
            )

        sections: list[ParsedSection] = []
        warnings: list[str] = []
        extracted_chars = 0
        retained_chars = 0
        nonempty_cells = 0
        formula_fallbacks = 0
        truncated = False

        for formula_sheet in visible_sheets:
            cached_sheet = cached_book[formula_sheet.title]
            max_column = max(
                int(formula_sheet.max_column or 0),
                int(cached_sheet.max_column or 0),
            )
            max_row = max(
                int(formula_sheet.max_row or 0),
                int(cached_sheet.max_row or 0),
            )
            if max_column > MAX_XLSX_COLUMNS:
                raise LocalDocumentParseError(
                    "xlsx_column_limit_exceeded",
                    "XLSX 单个工作表超过 200 列，请精简列或改用 Data X 分析。",
                )
            if max_row > MAX_XLSX_ROW_SPAN:
                raise LocalDocumentParseError(
                    "xlsx_cell_limit_exceeded",
                    "XLSX 的有效单元格或行跨度超过 100,000，请拆分工作簿或改用 Data X 分析。",
                )
            if max_column < 1 or max_row < 1:
                continue

            section_lines: list[str] = []
            section_chars = 0
            section_first_row: int | None = None
            section_last_row: int | None = None
            section_min_column: int | None = None
            section_max_column: int | None = None

            formula_rows = formula_sheet.iter_rows(
                min_row=1,
                max_row=max_row,
                min_col=1,
                max_col=max_column,
            )
            cached_rows = cached_sheet.iter_rows(
                min_row=1,
                max_row=max_row,
                min_col=1,
                max_col=max_column,
            )
            for row_number, (formula_row, cached_row) in enumerate(
                zip_longest(formula_rows, cached_rows, fillvalue=()),
                start=1,
            ):
                rendered_cells: list[str] = []
                row_min_column: int | None = None
                row_max_column: int | None = None
                for column_number in range(1, max_column + 1):
                    formula_cell = (
                        formula_row[column_number - 1]
                        if column_number <= len(formula_row)
                        else None
                    )
                    cached_cell = (
                        cached_row[column_number - 1]
                        if column_number <= len(cached_row)
                        else None
                    )
                    formula_value = getattr(formula_cell, "value", None)
                    cached_value = getattr(cached_cell, "value", None)
                    if not _xlsx_cell_is_present(formula_value, cached_value):
                        continue
                    nonempty_cells += 1
                    if nonempty_cells > MAX_XLSX_NONEMPTY_CELLS:
                        raise LocalDocumentParseError(
                            "xlsx_cell_limit_exceeded",
                            "XLSX 超过 100,000 个非空单元格，请拆分工作簿或改用 Data X 分析。",
                        )
                    is_formula = getattr(formula_cell, "data_type", None) == "f"
                    if is_formula and cached_value is None:
                        value = formula_value
                        formula_fallbacks += 1
                    elif is_formula:
                        value = cached_value
                    else:
                        value = formula_value
                    coordinate = f"{get_column_letter(column_number)}{row_number}"
                    rendered_cells.append(
                        f"{coordinate}: {_render_xlsx_cell_value(value)}"
                    )
                    row_min_column = (
                        column_number
                        if row_min_column is None
                        else min(row_min_column, column_number)
                    )
                    row_max_column = (
                        column_number
                        if row_max_column is None
                        else max(row_max_column, column_number)
                    )

                if not rendered_cells:
                    continue
                rendered_row = " | ".join(rendered_cells) + "\n"
                extracted_chars += len(rendered_row)
                remaining = MAX_EXTRACTED_CHARACTERS - retained_chars
                if remaining <= 0:
                    truncated = True
                    continue
                retained_row = rendered_row[:remaining]
                if len(retained_row) < len(rendered_row):
                    truncated = True
                if (
                    section_lines
                    and section_chars + len(retained_row) > MAX_SECTION_CHARACTERS
                ):
                    _append_xlsx_section(
                        sections,
                        sheet=formula_sheet.title,
                        lines=section_lines,
                        first_row=section_first_row,
                        last_row=section_last_row,
                        min_column=section_min_column,
                        max_column=section_max_column,
                    )
                    section_lines = []
                    section_chars = 0
                    section_first_row = None
                    section_last_row = None
                    section_min_column = None
                    section_max_column = None
                section_lines.append(retained_row)
                section_chars += len(retained_row)
                retained_chars += len(retained_row)
                section_first_row = row_number if section_first_row is None else section_first_row
                section_last_row = row_number
                section_min_column = (
                    row_min_column
                    if section_min_column is None
                    else min(section_min_column, row_min_column or section_min_column)
                )
                section_max_column = (
                    row_max_column
                    if section_max_column is None
                    else max(section_max_column, row_max_column or section_max_column)
                )

            if section_lines:
                _append_xlsx_section(
                    sections,
                    sheet=formula_sheet.title,
                    lines=section_lines,
                    first_row=section_first_row,
                    last_row=section_last_row,
                    min_column=section_min_column,
                    max_column=section_max_column,
                )

        if hidden_sheets:
            sample = "、".join(hidden_sheets[:8])
            suffix = " 等" if len(hidden_sheets) > 8 else ""
            warnings.append(f"已忽略隐藏工作表：{sample}{suffix}。")
        if formula_fallbacks:
            warnings.append(
                f"有 {formula_fallbacks} 个公式没有缓存结果，已保留公式文本且未执行公式。"
            )
        if truncated:
            warnings.append(
                "内容超过 500,000 字符，预览已安全截断；发送时将按模型上下文继续裁剪。"
            )
        if not sections:
            raise LocalDocumentParseError(
                "file_has_no_readable_text",
                "XLSX 的可见工作表中没有可读取的单元格。",
            )
        return ParsedDocument(
            format="xlsx",
            title=title or None,
            sections=tuple(sections),
            warnings=tuple(warnings),
            extracted_chars=extracted_chars,
            truncated=truncated,
        )
    except LocalDocumentParseError:
        raise
    except (InvalidFileException, OSError, ValueError, KeyError, TypeError) as exc:
        raise LocalDocumentParseError(
            "xlsx_parse_failed",
            "XLSX 无法安全读取，请确认文件未损坏并重新导出后再试。",
        ) from exc
    except Exception as exc:
        raise LocalDocumentParseError(
            "xlsx_parse_failed",
            "XLSX 解析意外失败，请重新导出或拆分工作簿后再试。",
        ) from exc
    finally:
        if cached_book is not None:
            cached_book.close()
        if formula_book is not None:
            formula_book.close()


def _xlsx_cell_is_present(formula_value: object, cached_value: object) -> bool:
    return any(value is not None and value != "" for value in (formula_value, cached_value))


def _render_xlsx_cell_value(value: object) -> str:
    if isinstance(value, datetime):
        rendered = value.isoformat(sep=" ")
    elif isinstance(value, (date, time)):
        rendered = value.isoformat()
    elif isinstance(value, timedelta):
        rendered = str(value)
    elif isinstance(value, bool):
        rendered = "true" if value else "false"
    else:
        rendered = str(value)
    return (
        rendered.replace("\\", "\\\\")
        .replace("|", "\\|")
        .replace("\r", " ")
        .replace("\n", "\\n")
    )


def _append_xlsx_section(
    sections: list[ParsedSection],
    *,
    sheet: str,
    lines: list[str],
    first_row: int | None,
    last_row: int | None,
    min_column: int | None,
    max_column: int | None,
) -> None:
    if not lines or None in {first_row, last_row, min_column, max_column}:
        return
    sections.append(
        ParsedSection(
            text="".join(lines).rstrip("\n"),
            sheet=sheet,
            row_range=(
                f"{get_column_letter(int(min_column))}{int(first_row)}:"
                f"{get_column_letter(int(max_column))}{int(last_row)}"
            ),
        )
    )


class _SafeHTMLTextExtractor(HTMLParser):
    _BLOCKED = {
        "script", "style", "template", "form", "iframe", "object", "embed",
        "svg", "math", "noscript",
    }
    _BLOCKS = {
        "p", "div", "section", "article", "main", "aside", "header", "footer",
        "li", "dt", "dd", "tr", "blockquote", "pre", "br", "hr",
    }

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.blocked_depth = 0
        self.buffer: list[str] = []
        self.sections: list[tuple[str, tuple[str, ...]]] = []
        self.headings: list[str] = []
        self.heading_level: int | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        del attrs
        clean = tag.lower()
        if self.blocked_depth:
            self.blocked_depth += 1
            return
        if clean in self._BLOCKED:
            self._flush()
            self.blocked_depth = 1
            return
        if clean in self._BLOCKS or _is_heading(clean):
            self._flush()
        if _is_heading(clean):
            self.heading_level = int(clean[1])

    def handle_endtag(self, tag: str) -> None:
        clean = tag.lower()
        if self.blocked_depth:
            self.blocked_depth -= 1
            return
        if _is_heading(clean):
            heading_text = " ".join("".join(self.buffer).split())
            if heading_text and self.heading_level is not None:
                self.headings = self.headings[: self.heading_level - 1]
                self.headings.append(heading_text)
            self._flush()
            self.heading_level = None
            return
        if clean in self._BLOCKS:
            self._flush()

    def handle_data(self, data: str) -> None:
        if not self.blocked_depth:
            self.buffer.append(data)

    def close(self) -> None:
        super().close()
        self._flush()

    def _flush(self) -> None:
        text = " ".join("".join(self.buffer).split())
        self.buffer = []
        if text:
            path = tuple(self.headings)
            if self.heading_level is not None:
                path = (*path[: self.heading_level - 1], text)
            self.sections.append((text, path))


def _is_heading(tag: str) -> bool:
    return len(tag) == 2 and tag[0] == "h" and tag[1] in "123456"


def _parse_html_text(path: Path, *, title: str | None) -> ParsedDocument:
    source_text = _read_utf8_text(path)
    extractor = _SafeHTMLTextExtractor()
    try:
        extractor.feed(source_text)
        extractor.close()
    except Exception as exc:
        raise LocalDocumentParseError(
            "file_parse_failed",
            "HTML 结构无法安全读取，请修复后重试。",
        ) from exc

    sections: list[ParsedSection] = []
    retained = 0
    truncated = False
    for text, heading_path in extractor.sections:
        remaining = MAX_EXTRACTED_CHARACTERS - retained
        if remaining <= 0:
            truncated = True
            break
        kept = text[:remaining]
        truncated = truncated or len(kept) < len(text)
        if kept:
            sections.append(
                ParsedSection(
                    text=kept,
                    heading_path=heading_path or None,
                )
            )
            retained += len(kept)
    if not sections:
        raise LocalDocumentParseError(
            "file_has_no_readable_text",
            "HTML 中没有可读取的安全正文。",
        )
    warnings = (
        ("HTML 中的脚本、样式、表单、嵌入对象和远程资源不会被执行或显示。",)
        + (("内容超过 500,000 字符，预览已安全截断；发送时将按模型上下文继续裁剪。",) if truncated else ())
    )
    return ParsedDocument(
        format="html",
        title=title or None,
        sections=tuple(sections),
        warnings=warnings,
        extracted_chars=len(source_text),
        truncated=truncated,
    )


def _parse_subtitle_text(
    path: Path,
    *,
    format_id: str,
    title: str | None,
) -> ParsedDocument:
    source_text = _read_utf8_text(path)
    sections: list[ParsedSection] = []
    retained = 0
    truncated = False
    for block, start_offset in _iter_subtitle_blocks(source_text):
        block_lines = block.splitlines()
        start_line = source_text.count("\n", 0, start_offset) + 1
        timeline = next((line.strip() for line in block_lines if "-->" in line), None)
        if timeline is None:
            continue
        remaining = MAX_EXTRACTED_CHARACTERS - retained
        if remaining <= 0:
            truncated = True
            break
        kept = block[:remaining]
        truncated = truncated or len(kept) < len(block)
        if kept.strip():
            sections.append(
                ParsedSection(
                    text=kept,
                    line_range=f"{start_line}-{start_line + len(block_lines) - 1}",
                    time_range=timeline,
                )
            )
            retained += len(kept)
    if not sections:
        raise LocalDocumentParseError(
            "file_has_no_readable_text",
            f"{format_id.upper()} 中没有可读取的字幕段落。",
        )
    warnings = (
        ("内容超过 500,000 字符，预览已安全截断；发送时将按模型上下文继续裁剪。",)
        if truncated
        else ()
    )
    return ParsedDocument(
        format=format_id,
        title=title or None,
        sections=tuple(sections),
        warnings=warnings,
        extracted_chars=len(source_text),
        truncated=truncated,
    )


def _iter_subtitle_blocks(text: str) -> list[tuple[str, int]]:
    """Return non-empty subtitle blocks with exact offsets in the source."""

    blocks: list[tuple[str, int]] = []
    cursor = 0
    for separator in re.finditer(r"(?:\r?\n[ \t]*){2,}", text):
        block = text[cursor : separator.start()]
        if block.strip():
            blocks.append((block, cursor))
        cursor = separator.end()
    block = text[cursor:]
    if block.strip():
        blocks.append((block, cursor))
    return blocks


def _parse_text_pdf(
    path: Path,
    *,
    title: str | None,
    timeout_seconds: float = PDF_PARSE_TIMEOUT_SECONDS,
    worker_target: PdfWorkerTarget | None = None,
) -> ParsedDocument:
    """Extract PDF text in a killable process with bounded output."""

    if timeout_seconds <= 0:
        raise ValueError("timeout_seconds must be positive")
    target = worker_target or _pdf_text_worker
    context = multiprocessing.get_context("spawn")
    receiver, sender = context.Pipe(duplex=False)
    worker = context.Process(
        target=target,
        args=(
            str(path),
            sender,
            MAX_PDF_PAGE_CHARACTERS,
            MAX_EXTRACTED_CHARACTERS,
            timeout_seconds,
        ),
        daemon=True,
    )
    try:
        worker.start()
    except Exception as exc:
        receiver.close()
        sender.close()
        raise LocalDocumentParseError(
            "pdf_parse_failed",
            "PDF 无法安全解析，请确认文件未损坏、未加密后重试。",
        ) from exc
    sender.close()

    message: tuple[str, Any] | None = None
    try:
        if not receiver.poll(timeout_seconds):
            _stop_pdf_worker(worker)
            raise LocalDocumentParseError(
                "pdf_parse_timeout",
                "PDF 文字提取超时。请拆分文档，或改用资料库视觉流水线处理。",
            )
        try:
            message = receiver.recv()
        except (EOFError, OSError) as exc:
            worker.join(timeout=0.5)
            if _pdf_worker_was_resource_limited(worker.exitcode):
                raise LocalDocumentParseError(
                    "pdf_parse_resource_limit",
                    "PDF 文字提取超过安全资源限制。请拆分或精简文档后再试。",
                ) from exc
            raise LocalDocumentParseError(
                "pdf_parse_failed",
                "PDF 文字提取进程意外结束，请重新导出文档后再试。",
            ) from exc
    finally:
        receiver.close()
        worker.join(timeout=0.5)
        if worker.is_alive():
            _stop_pdf_worker(worker)
        worker.close()

    if not isinstance(message, tuple) or len(message) != 2:
        raise LocalDocumentParseError(
            "pdf_parse_failed",
            "PDF 文字提取结果无效，请重新导出文档后再试。",
        )
    kind, payload = message
    if kind == "error" and isinstance(payload, dict):
        raise LocalDocumentParseError(
            str(payload.get("code") or "pdf_parse_failed"),
            str(
                payload.get("message")
                or "PDF 无法安全解析，请重新导出文档后再试。"
            ),
        )
    if kind != "ok" or not isinstance(payload, dict):
        raise LocalDocumentParseError(
            "pdf_parse_failed",
            "PDF 文字提取结果无效，请重新导出文档后再试。",
        )
    try:
        return ParsedDocument.model_validate({**payload, "title": title or None})
    except Exception as exc:
        raise LocalDocumentParseError(
            "pdf_parse_failed",
            "PDF 文字提取结果无效，请重新导出文档后再试。",
        ) from exc


def _stop_pdf_worker(worker: multiprocessing.Process) -> None:
    if worker.is_alive():
        worker.terminate()
        worker.join(timeout=1)
    if worker.is_alive() and hasattr(worker, "kill"):
        worker.kill()
        worker.join(timeout=1)


def _pdf_worker_was_resource_limited(exit_code: int | None) -> bool:
    """Recognize OS terminations caused by the worker's hard resource guard."""

    if exit_code is None or exit_code >= 0:
        return False
    resource_signals = {
        int(value)
        for name in ("SIGABRT", "SIGKILL", "SIGXCPU")
        if (value := getattr(signal, name, None)) is not None
    }
    return -exit_code in resource_signals


def _pdf_text_worker(
    path: str,
    sender: Connection,
    max_page_characters: int,
    max_total_characters: int,
    timeout_seconds: float,
) -> None:
    """Child-process worker; only returns bounded primitive data."""

    try:
        import PyPDF2  # type: ignore[import-not-found]

        # Import the parser before lowering the address-space ceiling. Spawned
        # workers otherwise need enough headroom to load the dependency itself.
        _apply_pdf_worker_resource_limits(timeout_seconds)
        with Path(path).open("rb") as source:
            reader = PyPDF2.PdfReader(source, strict=True)
            sections: list[dict[str, object]] = []
            empty_pages: list[int] = []
            retained_characters = 0
            extracted_characters = 0
            truncated = False
            for page_number, page in enumerate(reader.pages, start=1):
                text = (page.extract_text() or "").strip()
                if len(text) > max_page_characters:
                    _send_pdf_worker_error(
                        sender,
                        "pdf_parse_resource_limit",
                        "PDF 单页可提取文字过多。请拆分或精简文档后再试。",
                    )
                    return
                extracted_characters += len(text)
                if not text:
                    empty_pages.append(page_number)
                    continue
                remaining = max_total_characters - retained_characters
                if remaining <= 0:
                    truncated = True
                    break
                retained = text[:remaining]
                sections.append({"text": retained, "page": page_number})
                retained_characters += len(retained)
                if len(retained) < len(text):
                    truncated = True
                    break

        if not sections:
            _send_pdf_worker_error(
                sender,
                "scanned_pdf_requires_ocr",
                "该 PDF 没有可提取文字，可能是扫描件。请改用资料库视觉流水线，或在后续显式确认 OCR 费用后处理。",
            )
            return

        warnings: list[str] = []
        if empty_pages:
            sample = "、".join(str(page) for page in empty_pages[:8])
            suffix = " 等" if len(empty_pages) > 8 else ""
            warnings.append(
                f"第 {sample}{suffix} 页未提取到文字，可能包含扫描图像。"
            )
        if truncated:
            warnings.append(
                "内容超过 500,000 字符，预览已安全截断；发送时将按模型上下文继续裁剪。"
            )
        sender.send(
            (
                "ok",
                {
                    "format": "pdf",
                    "sections": sections,
                    "warnings": warnings,
                    "extracted_chars": extracted_characters,
                    "truncated": truncated,
                },
            )
        )
    except MemoryError:
        _send_pdf_worker_error(
            sender,
            "pdf_parse_resource_limit",
            "PDF 文字提取超过安全资源限制。请拆分或精简文档后再试。",
        )
    except Exception:
        _send_pdf_worker_error(
            sender,
            "pdf_parse_failed",
            "PDF 无法安全解析，请确认文件未损坏、未加密后重试。",
        )
    finally:
        sender.close()


def _send_pdf_worker_error(
    sender: Connection,
    code: str,
    message: str,
) -> None:
    try:
        sender.send(("error", {"code": code, "message": message}))
    except (BrokenPipeError, EOFError, OSError):
        pass


def _apply_pdf_worker_resource_limits(timeout_seconds: float) -> None:
    """Apply hard OS limits where available; Windows relies on process kill."""

    try:
        import resource

        current_virtual_memory = _current_virtual_memory_bytes()
        address_space_limit = _pdf_worker_address_space_limit(
            current_virtual_memory
        )
        # ``spawn`` imports this package before entering the target. Native
        # dependencies can reserve a large virtual address range during that
        # import while using little resident memory. A 512 MiB absolute ceiling
        # can therefore sit below the existing baseline and make the next libc
        # TLS allocation abort. Keep the same fixed 512 MiB growth budget above
        # the measured baseline instead.
        resource.setrlimit(
            resource.RLIMIT_AS,
            (address_space_limit, address_space_limit),
        )
        cpu_seconds = max(1, int(timeout_seconds) + 1)
        resource.setrlimit(resource.RLIMIT_CPU, (cpu_seconds, cpu_seconds + 1))
    except (ImportError, OSError, ValueError):
        return


def _pdf_worker_address_space_limit(current_virtual_memory: int | None) -> int:
    """Keep a finite PDF allocation budget above the spawned worker baseline."""

    baseline = max(0, current_virtual_memory or 0)
    return baseline + PDF_PARSE_WORKER_MEMORY_BYTES


def _current_virtual_memory_bytes() -> int | None:
    """Return the Linux process address-space baseline without extra imports."""

    try:
        pages = int(Path("/proc/self/statm").read_text(encoding="ascii").split()[0])
        page_size = int(os.sysconf("SC_PAGE_SIZE"))
    except (IndexError, OSError, TypeError, ValueError):
        return None
    if pages <= 0 or page_size <= 0:
        return None
    return pages * page_size
